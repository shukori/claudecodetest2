# -*- coding: utf-8 -*-
"""
在宅訪問調剤薬局 事業者集計ビルダー（A案・改訂版）
 出力1: 福岡県_事業者集計.xlsx      （基本一覧×施設基準をコード結合＝法人名寄せ＋在宅フラグ）
 出力2: 九州8県_在宅薬局_事業者集計.xlsx（福岡含む九州全8県。福岡=法人ベース／他7県=屋号ベースのハイブリッド名寄せ）
 設計: 集計はCOUNTIFSでローデータ参照／名寄せ辞書・名寄せ履歴・要確認修正シートで監査可能
"""
import openpyxl, re, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, OrderedDict

SRC = "/root/.claude/uploads/5f0eed8e-ea94-5c75-b8ad-040dd274105a"
OUT = "/home/user/ClaudeCodetest2"

# ================= 名寄せ辞書（改訂） =================
# (ルールID, 正規化事業者名, タイプ, 親会社/メモ, 信頼度, [マッチ語(法人 or 屋号に部分一致)])
RULES = [
 ("R01","アインホールディングス（アイン薬局）","大手全国","東証PRM・国内最大手","高",["アインファーマシーズ","アイン薬局","アインメディオ"]),
 ("R02","総合メディカル（そうごう薬局）","大手全国","SOGO MEDICAL・福岡発祥の全国大手","高",["総合メディカル","そうごう薬局","ソウゴウ薬局"]),
 ("R03","日本調剤","大手全国","東証PRM","高",["日本調剤"]),
 ("R04","クオール（クオール薬局）","大手全国","東証PRM","高",["クオール"]),
 ("R05","メディカルシステムネットワーク（なの花薬局）","大手全国","札幌発祥・なの花・九州はなの花九州／さくら薬局屋号も同社系が多い","高",["なの花","メディカルシステムネットワーク"]),
 ("R06","クラフト（さくら薬局グループ）","大手全国","※「さくら薬局」屋号は九州ではなの花九州系が大半→屋号では集約せず法人名でのみ判定","中",["クラフト株式会社","クラフト"]),
 ("R07","HYUGA PRIMARY CARE（きらり薬局）","在宅特化・上場","東証GRT・分析対象本命","高",["ＨＹＵＧＡ","HYUGA","きらり薬局"]),
 ("R08","ウエルシアHD（イオン系）","ドラッグ大手","調剤併設","高",["ウエルシア","イオンウエルシア"]),
 ("R09","ツルハHD","ドラッグ大手","調剤併設","高",["ツルハ","くすりの福太郎"]),
 ("R10","マツキヨココカラ&Co","ドラッグ大手","調剤併設","高",["ココカラファイン","マツモトキヨシ"]),
 ("R11","スギHD（スギ薬局）","ドラッグ大手","調剤併設","高",["スギ薬局","スギホールディングス"]),
 # --- 九州・福岡 地場大手/中堅 ---
 ("R20","大賀薬局","地場大手(福岡)","福岡地盤の有力チェーン","高",["大賀薬局"]),
 ("R21","新生堂薬局","地場大手(福岡)","福岡・ドラッグ＋調剤","高",["新生堂"]),
 ("R22","サンキュードラッグ","地場(北九州)","北九州地盤ドラッグ","高",["サンキュードラッグ"]),
 ("R23","タカラ薬局（大蔵商事/タカラ）","地場(福岡)","福岡地盤","中",["タカラ薬局"]),
 ("R24","溝上薬局","地場大手(福岡)","福岡・九州で展開","高",["溝上薬局"]),
 ("R25","タケシタ調剤薬局","地場(福岡・熊本)","","高",["タケシタ"]),
 ("R26","コスモス薬品","ドラッグ(九州発)","ディスカウントドラッグコスモス","高",["コスモス薬品","ディスカウントドラッグコスモス"]),
 ("R27","ドラッグストアモリ","ドラッグ(九州)","","高",["ドラッグストアモリ"]),
 ("R28","ドラッグイレブン","ドラッグ(九州)","JR九州系","中",["ドラッグイレブン"]),
 ("R29","大信薬局","地場(福岡)","在宅・施設に注力","中",["大信薬局"]),
 ("R30","ニック（ニック調剤薬局）","地場(九州)","","中",["ニック"]),
 ("R31","アガペ","地場(福岡)","","中",["アガペ"]),
 ("R32","ミズ","地場(福岡)","","中",["株式会社ミズ","ミズ薬局"]),
 ("R33","タカサキ（高崎）薬局","地場","","中",["タカサキ","高崎薬局"]),
 # --- 要確認→解消済（下記 03_要確認修正 参照）---
 ("R40","すこやか薬局（沖縄地盤）","地場(沖縄)","沖縄45店の単一チェーン。熊本/長崎の同名は別法人の可能性（屋号一致のため混入注意）","中",["すこやか薬局"]),
]

# 要確認の解消記録（03シートに出力）
CORRECTIONS = [
 ("さくら薬局",
  "福岡58件の法人内訳＝なの花九州41／クラフト1／有限会社さくら薬局3／他独立。屋号『さくら薬局』はなの花九州系（R05）が大半でクラフトは僅少。",
  "R06『クラフト（さくら薬局）』へ屋号一致で全件誤集約",
  "R06のエイリアスから『さくら薬局』を削除。福岡は法人で正しく分解（なの花九州→R05等）、非福岡は屋号baseで独立計上（同名異法人は分離不可と明記）",
  "福岡の さくら薬局≈41店が クラフト→メディカルSN(なの花) へ是正","高"),
 ("さくら調剤薬局",
  "福岡5件で法人4社混在（さくら調剤/おおじま調剤/中商/朝倉ファーマシー）。旧ルールは『あさくら調剤薬局』も誤マッチ。",
  "R41『さくら調剤薬局』を単一事業者として集約",
  "R41を廃止。福岡は法人で分解。非福岡は屋号base『さくら調剤』で計上（同名複数法人の可能性を注記）。『あさくら』は別baseで自動分離",
  "単一チェーンの誤認を解消（実態は複数法人）","高"),
 ("すこやか薬局",
  "全47件中 沖縄45・熊本1・長崎1。沖縄45は屋号＋店名で単一チェーンと判断。",
  "R40 要確認（同名複数の可能性）",
  "R40『すこやか薬局（沖縄地盤）』として確定（信頼度中）。熊本/長崎の同名2件は別法人の可能性ありと注記し屋号一致で混入する点を明示",
  "沖縄チェーンとして確定、2件の他県同名は誤差","中"),
 ("（全般）屋号ベースの限界",
  "非福岡7県は施設基準ファイルに開設者法人列が無く屋号でしか名寄せできない。",
  "－",
  "同名異法人（ひまわり/あおぞら/さくら調剤等）は分離不可。大手は法人裏付けのある福岡＋全国チェーン辞書で担保。生データに屋号コアを残し追跡可能に",
  "中小・独立カテゴリは過小/過大の双方向誤差を含みうる","中"),
]

def clean_corp(kaisetsu):
    if not kaisetsu: return ""
    s=str(kaisetsu)
    s=re.split(r'(代表取締役|代表者|取締役|理事長|理事|院長|管理者|会長|社長|代表社員|業務執行社員|開設者)',s)[0]
    return s.replace('　',' ').strip()

def base_yago(name):
    if not name: return ""
    s=str(name).replace('　','').replace(' ','')
    for t in ['株式会社','有限会社','合同会社','合資会社','医療法人社団','医療法人','一般社団法人','社会福祉法人','公益財団法人']:
        s=s.replace(t,'')
    m=re.search(r'^(.*?(薬局|ファーマシー|ドラッグ|薬店|調剤|ファーマ))',s)
    core=m.group(1) if m else s[:10]
    if core in ('薬局','調剤','保険薬局') or len(core)<2:
        core=s[:10]
    return core

def match_rule(text):
    s=str(text or "").replace('　','').replace(' ','')
    for rid,name,typ,memo,conf,aliases in RULES:
        for a in aliases:
            if a.replace('　','').replace(' ','') in s:
                return rid,name,typ,conf
    return None

def norm_code(s): return str(s).strip().replace(',','')

# ================= スタイル =================
HEAD=Font(bold=True,color="FFFFFF",size=11,name="IPAGothic"); HFILL=PatternFill("solid",fgColor="2E6DA4")
GFILL=PatternFill("solid",fgColor="E3F2E9"); YFILL=PatternFill("solid",fgColor="FFF2CC")
RFILL=PatternFill("solid",fgColor="FCE4E4")
TITLE=Font(bold=True,size=14,name="IPAGothic"); NORM=Font(size=10,name="IPAGothic"); BOLD=Font(bold=True,size=10,name="IPAGothic")
thin=Side(style="thin",color="BBBBBB"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top"); CEN=Alignment(horizontal="center",vertical="center")
def style_header(ws,row,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(row=row,column=c); cell.font=HEAD; cell.fill=HFILL; cell.alignment=CEN; cell.border=BORD
def link_font(): return Font(size=10,name="IPAGothic",color="0563C1",underline="single")

# ================= ファイル定義 =================
FUKUOKA_BASIC="c253f318-r8_06_fukuoka_yakkyoku_02.xlsx"
SHISETSU=OrderedDict([
 ("福岡県","23342771-r8_05_shisetsu_fukuoka_yakkyoku_02.xlsx"),
 ("佐賀県","3b941e75-r8_05_shisetsu_saga_yakkyoku_02.xlsx"),
 ("長崎県","e0bd4eb7-r8_05_shisetsu_nagasaki_yakkyoku_02.xlsx"),
 ("熊本県","a0c45bc5-r8_05_shisetsu_kumamoto_yakkyoku_02.xlsx"),
 ("大分県","19405bbf-r8_05_shisetsu_ooita_yakkyoku_02.xlsx"),
 ("宮崎県","da1c0ae4-r8_05_shisetsu_miyazaki_yakkyoku_02.xlsx"),
 ("鹿児島県","4ad91e46-r8_05_shisetsu_kagoshima_yakkyoku_02.xlsx"),
 ("沖縄県","0655d75f-r8_05_shisetsu_okinawa_yakkyoku_02.xlsx"),
])
SRC_URL="https://kouseikyoku.mhlw.go.jp/kyushu/gyomu/gyomu/hoken_kikan/index_00004.html"
TAISEI={"在宅薬学総合体制加算１","在宅薬学総合体制加算２"}; HOMEV="在宅患者訪問薬剤管理指導料"

# ================= パース =================
def parse_fukuoka_basic():
    """福岡基本一覧 → code→(屋号,開設者法人)"""
    wb=openpyxl.load_workbook(os.path.join(SRC,FUKUOKA_BASIC),read_only=True); ws=wb["Sheet1"]
    out=OrderedDict()
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or not str(row[0]).strip().isdigit(): continue
        if row[1] is None or not re.match(r'^\d{2,3},\d{3},\d$',str(row[1]).strip()): continue
        out[norm_code(row[1])]=dict(code=norm_code(row[1]),name=row[2],addr=row[3],corp_raw=clean_corp(row[5]))
    wb.close(); return out

def parse_shisetsu_one(pref,f):
    """施設基準1県 → code→{name,addr,flags}"""
    wb=openpyxl.load_workbook(os.path.join(SRC,f),read_only=True); ws=wb["Sheet1"]
    ph=OrderedDict()
    for row in ws.iter_rows(min_row=5,values_only=True):
        if row[0] is None or not str(row[0]).strip().isdigit(): continue
        code=row[4]
        if code is None: continue
        c=norm_code(code)
        if c not in ph: ph[c]=dict(pref=pref,code=c,name=row[7],addr=row[9],flags=set())
        if row[13]: ph[c]["flags"].add(row[13])
    wb.close(); return ph

def flags_cols(fl):
    t1=1 if "在宅薬学総合体制加算１" in fl else 0
    t2=1 if "在宅薬学総合体制加算２" in fl else 0
    return t1,t2,(1 if t1 or t2 else 0),(1 if HOMEV in fl else 0)

# ================= 共通シート =================
def add_dict_sheet(wb):
    ws=wb.create_sheet("01_名寄せ辞書")
    ws["A1"]="名寄せ辞書（大手・主要チェーン／改訂版）"; ws["A1"].font=TITLE
    ws["A2"]="上から順に評価し最初に一致したルールを採用（first match wins）。福岡は開設者法人、非福岡は屋号に対して照合。未一致は『中小・独立』。"; ws["A2"].font=NORM
    heads=["ルールID","正規化事業者名","タイプ","親会社・メモ","信頼度","マッチ語（部分一致）"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads)); r=5
    for rid,name,typ,memo,conf,aliases in RULES:
        for j,v in enumerate([rid,name,typ,memo,conf,"；".join(aliases)],1):
            c=ws.cell(row=r,column=j,value=v); c.font=NORM; c.border=BORD; c.alignment=WRAP
        if conf=="要確認":
            for j in range(1,7): ws.cell(row=r,column=j).fill=YFILL
        r+=1
    for j,w in enumerate([10,36,16,40,8,40],1): ws.column_dimensions[get_column_letter(j)].width=w

def add_history_sheet(wb,hist,raw_label):
    ws=wb.create_sheet("02_名寄せ履歴")
    ws["A1"]="名寄せ履歴（監査用：原名称→適用ルール→正規化事業者）"; ws["A1"].font=TITLE
    ws["A2"]=f"原データの『{raw_label}』を一意化し、どのルールで・どの事業者へ寄せたかを全件表示。件数合計はローデータ行数と一致。"; ws["A2"].font=NORM
    heads=[f"原・{raw_label}","名寄せ方式","適用ルールID","正規化事業者名","タイプ","信頼度","該当店舗数"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads))
    hist_sorted=sorted(hist,key=lambda x:(x["rid"]=="-",-x["cnt"])); r=5
    for h in hist_sorted:
        for j,v in enumerate([h["raw"],h["via"],h["rid"],h["corp"],h["type"],h["conf"],h["cnt"]],1):
            c=ws.cell(row=r,column=j,value=v); c.font=NORM; c.border=BORD
        if h["conf"]=="要確認":
            for j in range(1,8): ws.cell(row=r,column=j).fill=YFILL
        elif h["rid"]!="-":
            for j in range(1,8): ws.cell(row=r,column=j).fill=GFILL
        r+=1
    ws.cell(row=r,column=6,value="合計").font=BOLD; ws.cell(row=r,column=7,value=f"=SUM(G5:G{r-1})").font=BOLD
    for j,w in enumerate([34,10,11,34,16,9,11],1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A5"

def add_corrections_sheet(wb):
    ws=wb.create_sheet("03_要確認修正")
    ws["A1"]="要確認名寄せの解消記録（全域で適用）"; ws["A1"].font=TITLE
    ws["A2"]="汎用屋号・誤集約を実データ（特に福岡の法人内訳）で検証し、辞書を修正した履歴。"; ws["A2"].font=NORM
    heads=["対象","調査結果（根拠）","旧マッピング","新マッピング（適用済）","影響","信頼度"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads)); r=5
    for tgt,found,old,new,impact,conf in CORRECTIONS:
        for j,v in enumerate([tgt,found,old,new,impact,conf],1):
            c=ws.cell(row=r,column=j,value=v); c.font=NORM; c.border=BORD; c.alignment=WRAP
        for j in range(1,7): ws.cell(row=r,column=j).fill=RFILL
        r+=1
    for j,w in enumerate([16,46,28,52,30,8],1): ws.column_dimensions[get_column_letter(j)].width=w

def write_readme(wb,lines):
    ws=wb.create_sheet("00_README")
    for i,(t,f) in enumerate(lines,1): ws.cell(row=i,column=1,value=t).font=f
    ws.column_dimensions["A"].width=114

# ================= 集計シート（COUNTIF系） =================
print("logic loaded")

# ================= 行構築（共通） =================
def build_rows():
    """全8県の在宅薬局ローデータ行を構築。福岡は法人結合、ハイブリッド名寄せ。"""
    fk_basic=parse_fukuoka_basic()  # code->(屋号,法人)
    rows=[]; hist=OrderedDict()
    for pref,f in SHISETSU.items():
        ph=parse_shisetsu_one(pref,f)
        for c,d in ph.items():
            t1,t2,taisei,homev=flags_cols(d["flags"])
            if pref=="福岡県":
                corp_raw=fk_basic.get(c,{}).get("corp_raw","")
                via="法人" if corp_raw else "屋号"
                key=corp_raw if corp_raw else d["name"]
                m=match_rule(corp_raw if corp_raw else d["name"])
                fallback=corp_raw if corp_raw else base_yago(d["name"])
                hist_raw=corp_raw if corp_raw else base_yago(d["name"])
            else:
                via="屋号"; key=d["name"]; m=match_rule(d["name"])
                fallback=base_yago(d["name"]); hist_raw=base_yago(d["name"])
            if m: rid,corp,typ,conf=m
            else: rid,corp,typ,conf="-",fallback,"中小・独立","自動"
            rows.append(dict(pref=pref,code=c,name=d["name"],addr=d["addr"],
                corp=corp,rid=rid,type=typ,conf=conf,via=via,yago=base_yago(d["name"]),
                corp_raw=(fk_basic.get(c,{}).get("corp_raw","") if pref=="福岡県" else ""),
                t1=t1,t2=t2,taisei=taisei,homev=homev))
            hk=(hist_raw,via,rid,corp)
            if hk not in hist: hist[hk]=dict(raw=hist_raw,via=via,rid=rid,corp=corp,type=typ,conf=conf,cnt=0)
            hist[hk]["cnt"]+=1
    return rows,list(hist.values())

# ================= 出力2: 九州8県 =================
def build_kyushu(rows,hist):
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    prefs=list(SHISETSU.keys())
    write_readme(wb,[
     ("九州全体8県 在宅薬局 事業者集計（A案・改訂版）",TITLE),("",NORM),
     ("【元データ】令和8年5月1日現在 施設基準届出状況（保険薬局）／九州厚生局　※福岡は6月1日基本一覧と結合",BOLD),
     (f"  出典URL: {SRC_URL}",link_font()),
     ("  対象: 福岡・佐賀・長崎・熊本・大分・宮崎・鹿児島・沖縄（九州全8県）",NORM),("",NORM),
     ("【在宅の定義】",BOLD),
     ("  ◎主定義=在宅薬学総合体制加算1or2の届出（本気の在宅体制／集計の既定軸）",NORM),
     ("  ○広義  =在宅患者訪問薬剤管理指導料の届出（届出のみ・休眠含む／参考）",NORM),("",NORM),
     ("【名寄せ方針（ハイブリッド）】",BOLD),
     ("  福岡=基本一覧と保険薬局コードで結合し『開設者法人』で名寄せ（高精度）。",NORM),
     ("  非福岡7県=施設基準ファイルに法人列が無いため『屋号』で名寄せ。",NORM),
     ("  ローデータの『名寄せ方式』列で法人/屋号を明示。要確認名は 03_要確認修正 で全域解消済。",NORM),("",NORM),
     ("【シート】01_名寄せ辞書 / 02_名寄せ履歴 / 03_要確認修正 / 10_ローデータ / 20_集計_事業者別 / 21_集計_県別サマリ",BOLD),
     ("  集計値は全てCOUNTIFS数式でローデータ参照。ローデータ修正で自動更新。",NORM),
    ])
    add_dict_sheet(wb); add_history_sheet(wb,hist,"法人/屋号"); add_corrections_sheet(wb)

    # 10_ローデータ
    wd=wb.create_sheet("10_ローデータ")
    heads=["No","都道府県","保険薬局コード","薬局名称(屋号)","所在地","開設者法人(福岡のみ)","名寄せ方式","屋号コア",
           "在総1","在総2","在総(1or2)","在薬(訪問届出)","適用ルールID","正規化事業者名","事業者タイプ","信頼度"]
    for j,h in enumerate(heads,1): wd.cell(row=1,column=j,value=h)
    style_header(wd,1,len(heads))
    for i,r in enumerate(rows,1):
        vals=[i,r["pref"],r["code"],r["name"],r["addr"],r["corp_raw"],r["via"],r["yago"],
              r["t1"],r["t2"],r["taisei"],r["homev"],r["rid"],r["corp"],r["type"],r["conf"]]
        for j,v in enumerate(vals,1):
            c=wd.cell(row=i+1,column=j,value=v); c.font=NORM; c.border=BORD
    wd.freeze_panes="A2"
    for j,w in enumerate([5,9,13,28,34,26,10,20,7,7,10,12,10,30,14,9],1): wd.column_dimensions[get_column_letter(j)].width=w
    nrow=len(rows)+1; CORP="N"; PREF="B"; TAI="K"; HOM="L"

    # 20_集計_事業者別（在総ベース）
    ws=wb.create_sheet("20_集計_事業者別")
    ws["A1"]="九州8県 在宅薬局 事業者別 店舗数（在宅薬学総合体制加算ベース）"; ws["A1"].font=TITLE
    ws["A2"]="在総店舗数=COUNTIFS(事業者一致 & 在総=1)。広義(在薬)も併記。並びは(1)ルール一致=大手優先 (2)同グループ内で在総数降順 (3)ルール未一致は最後。"; ws["A2"].font=NORM
    heads=["順位","正規化事業者名","事業者タイプ","信頼度","在総薬局数","広義(在薬)薬局数","在総シェア"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads))
    agg=OrderedDict()
    for r in rows:
        k=r["corp"]
        if k not in agg: agg[k]=dict(type=r["type"],conf=r["conf"],rid=r["rid"],taisei=0)
        agg[k]["taisei"]+=r["taisei"]
    # ソート順序: (1)ルール一致（大手優先） (2)在総数降順 (3)同数時は信頼度（高>中>自動）
    conf_rank={"高":0,"中":1,"自動":2}
    order=sorted([(k,v) for k,v in agg.items() if v["taisei"]>0],
                 key=lambda kv:(kv[1]["rid"]=="-",-kv[1]["taisei"],conf_rank.get(kv[1]["conf"],3)))
    r=5
    for rank,(corp,info) in enumerate(order,1):
        ws.cell(row=r,column=1,value=rank); ws.cell(row=r,column=2,value=corp)
        ws.cell(row=r,column=3,value=info["type"]); ws.cell(row=r,column=4,value=info["conf"])
        ws.cell(row=r,column=5,value=f"=COUNTIFS('10_ローデータ'!{CORP}:{CORP},B{r},'10_ローデータ'!{TAI}:{TAI},1)")
        ws.cell(row=r,column=6,value=f"=COUNTIFS('10_ローデータ'!{CORP}:{CORP},B{r},'10_ローデータ'!{HOM}:{HOM},1)")
        ws.cell(row=r,column=7,value=f"=E{r}/$E${5+len(order)+1}"); ws.cell(row=r,column=7).number_format='0.0%'
        for c in range(1,8): ws.cell(row=r,column=c).font=NORM; ws.cell(row=r,column=c).border=BORD
        if info["rid"]!="-" and info["conf"] not in ("要確認",):
            for c in range(1,8): ws.cell(row=r,column=c).fill=GFILL
        r+=1
    ws.cell(row=r,column=2,value="合計（検算）").font=BOLD; ws.cell(row=r,column=5,value=f"=SUM(E5:E{r-1})").font=BOLD
    ws.cell(row=r,column=6,value=f"=SUM(F5:F{r-1})").font=BOLD
    ws.cell(row=r+1,column=2,value="在総薬局 総数（ローデータ）").font=NORM
    ws.cell(row=r+1,column=5,value=f"=COUNTIF('10_ローデータ'!{TAI}2:{TAI}{nrow},1)").font=NORM
    for j,w in enumerate([6,34,16,9,12,15,11],1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A5"

    # 21_県別サマリ
    ws=wb.create_sheet("21_集計_県別サマリ")
    ws["A1"]="県別サマリ（在宅薬局数・在宅実施率）"; ws["A1"].font=TITLE
    ws["A2"]="COUNTIFS参照。実施率=在総薬局数/全薬局数。"; ws["A2"].font=NORM
    heads=["都道府県","全薬局数","在総薬局数(本気)","在総実施率","広義(在薬)薬局数","広義実施率"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads)); r=5
    for pref in prefs:
        ws.cell(row=r,column=1,value=pref)
        ws.cell(row=r,column=2,value=f"=COUNTIF('10_ローデータ'!{PREF}:{PREF},A{r})")
        ws.cell(row=r,column=3,value=f"=COUNTIFS('10_ローデータ'!{PREF}:{PREF},A{r},'10_ローデータ'!{TAI}:{TAI},1)")
        ws.cell(row=r,column=4,value=f"=C{r}/B{r}"); ws.cell(row=r,column=4).number_format='0.0%'
        ws.cell(row=r,column=5,value=f"=COUNTIFS('10_ローデータ'!{PREF}:{PREF},A{r},'10_ローデータ'!{HOM}:{HOM},1)")
        ws.cell(row=r,column=6,value=f"=E{r}/B{r}"); ws.cell(row=r,column=6).number_format='0.0%'
        for c in range(1,7): ws.cell(row=r,column=c).font=NORM; ws.cell(row=r,column=c).border=BORD
        r+=1
    ws.cell(row=r,column=1,value="8県合計").font=BOLD
    for col,L in [(2,"B"),(3,"C"),(5,"E")]: ws.cell(row=r,column=col,value=f"=SUM({L}5:{L}{r-1})").font=BOLD
    ws.cell(row=r,column=4,value=f"=C{r}/B{r}").font=BOLD; ws.cell(row=r,column=4).number_format='0.0%'
    ws.cell(row=r,column=6,value=f"=E{r}/B{r}").font=BOLD; ws.cell(row=r,column=6).number_format='0.0%'
    for j,w in enumerate([12,12,16,12,16,12],1): ws.column_dimensions[get_column_letter(j)].width=w

    path=f"{OUT}/九州8県_在宅薬局_事業者集計.xlsx"; wb.save(path); return path,len(rows),len(order)

# ================= 出力1: 福岡（法人＋在宅） =================
def build_fukuoka(rows,hist):
    fk=[r for r in rows if r["pref"]=="福岡県"]
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    write_readme(wb,[
     ("福岡県 事業者集計（基本一覧×施設基準 結合・改訂版）",TITLE),("",NORM),
     ("【元データ】①保険薬局一覧(令和8年6月1日現在) ②施設基準届出状況(令和8年5月1日現在)／九州厚生局",BOLD),
     (f"  出典URL: {SRC_URL}",link_font()),
     ("  ①r8_06_fukuoka_yakkyoku_02.xlsx（開設者法人）× ②r8_05_shisetsu_fukuoka_yakkyoku_02.xlsx（在宅フラグ）を保険薬局コードで結合。",NORM),
     ("  結合率: 施設側2,943件のうち基本一覧と突合可能（在宅フラグ＋法人名を両取り）。",NORM),("",NORM),
     ("【在宅の定義】◎在宅薬学総合体制加算1or2 ／ ○在宅患者訪問薬剤管理指導料（広義）",BOLD),("",NORM),
     ("【名寄せ】開設者法人をキーに辞書(01)で正規化。法人があるため屋号より高精度（例: 屋号さくら薬局→法人なの花九州を正しく識別）。",BOLD),
     ("  要確認名は 03_要確認修正 で全域解消済（さくら薬局/さくら調剤/すこやか薬局）。",NORM),("",NORM),
     ("【シート】01_名寄せ辞書 / 02_名寄せ履歴 / 03_要確認修正 / 10_ローデータ / 20_集計_全薬局事業者別 / 21_集計_在宅事業者別",BOLD),
     ("  20=在宅フィルタ無しの全薬局シェア、21=在総ベースの在宅事業者ランキング。いずれもCOUNTIFS数式。",NORM),
    ])
    # 福岡専用の名寄せ履歴（福岡行から再構築：法人があれば法人、無ければ屋号コア）
    fk_hist=OrderedDict()
    for r in fk:
        raw=r["corp_raw"] if r["corp_raw"] else r["yago"]
        hk=(raw,r["via"],r["rid"],r["corp"])
        if hk not in fk_hist: fk_hist[hk]=dict(raw=raw,via=r["via"],rid=r["rid"],corp=r["corp"],type=r["type"],conf=r["conf"],cnt=0)
        fk_hist[hk]["cnt"]+=1
    add_dict_sheet(wb); add_history_sheet(wb,list(fk_hist.values()),"開設者法人/屋号"); add_corrections_sheet(wb)

    wd=wb.create_sheet("10_ローデータ")
    heads=["No","保険薬局コード","薬局名称(屋号)","所在地","開設者法人","適用ルールID","正規化事業者名","事業者タイプ","信頼度",
           "在総1","在総2","在総(1or2)","在薬(訪問届出)"]
    for j,h in enumerate(heads,1): wd.cell(row=1,column=j,value=h)
    style_header(wd,1,len(heads))
    for i,r in enumerate(fk,1):
        vals=[i,r["code"],r["name"],r["addr"],r["corp_raw"],r["rid"],r["corp"],r["type"],r["conf"],
              r["t1"],r["t2"],r["taisei"],r["homev"]]
        for j,v in enumerate(vals,1):
            c=wd.cell(row=i+1,column=j,value=v); c.font=NORM; c.border=BORD
    wd.freeze_panes="A2"
    for j,w in enumerate([5,13,28,36,30,10,30,14,9,7,7,10,12],1): wd.column_dimensions[get_column_letter(j)].width=w
    nrow=len(fk)+1; CORP="G"; TAI="L"; HOM="M"

    # 20 全薬局事業者別
    ws=wb.create_sheet("20_集計_全薬局事業者別")
    ws["A1"]="福岡県 事業者別 店舗数（全薬局・在宅フィルタ無）"; ws["A1"].font=TITLE
    ws["A2"]="店舗数=COUNTIF(正規化事業者)。並びは(1)ルール一致=大手優先 (2)同グループ内で店舗数降順 (3)ルール未一致は最後。"; ws["A2"].font=NORM
    heads=["順位","正規化事業者名","事業者タイプ","信頼度","店舗数","シェア"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads))
    agg=OrderedDict()
    for r in fk:
        k=r["corp"]
        if k not in agg: agg[k]=dict(type=r["type"],conf=r["conf"],rid=r["rid"],cnt=0,tai=0)
        agg[k]["cnt"]+=1; agg[k]["tai"]+=r["taisei"]
    conf_rank={"高":0,"中":1,"自動":2}
    order=sorted(agg.items(),key=lambda kv:(kv[1]["rid"]=="-",-kv[1]["cnt"],conf_rank.get(kv[1]["conf"],3))); r=5
    for rank,(corp,info) in enumerate(order,1):
        ws.cell(row=r,column=1,value=rank); ws.cell(row=r,column=2,value=corp)
        ws.cell(row=r,column=3,value=info["type"]); ws.cell(row=r,column=4,value=info["conf"])
        ws.cell(row=r,column=5,value=f"=COUNTIF('10_ローデータ'!{CORP}:{CORP},B{r})")
        ws.cell(row=r,column=6,value=f"=E{r}/$E${5+len(order)+1}"); ws.cell(row=r,column=6).number_format='0.0%'
        for c in range(1,7): ws.cell(row=r,column=c).font=NORM; ws.cell(row=r,column=c).border=BORD
        if info["rid"]!="-":
            for c in range(1,7): ws.cell(row=r,column=c).fill=GFILL
        r+=1
    ws.cell(row=r,column=2,value="合計（検算）").font=BOLD; ws.cell(row=r,column=5,value=f"=SUM(E5:E{r-1})").font=BOLD
    ws.cell(row=r+1,column=2,value="ローデータ行数").font=NORM
    ws.cell(row=r+1,column=5,value=f"=COUNTA('10_ローデータ'!B2:B{nrow})").font=NORM
    for j,w in enumerate([6,32,16,9,10,10],1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A5"

    # 21 在宅事業者別（在総）
    ws=wb.create_sheet("21_集計_在宅事業者別")
    ws["A1"]="福岡県 在宅薬局 事業者別 店舗数（在宅薬学総合体制加算ベース）"; ws["A1"].font=TITLE
    ws["A2"]="在総店舗数=COUNTIFS(事業者一致 & 在総=1)。法人名寄せにより高精度。並びは(1)ルール一致=大手優先 (2)同グループ内で在総数降順 (3)ルール未一致は最後。"; ws["A2"].font=NORM
    heads=["順位","正規化事業者名","事業者タイプ","信頼度","在総薬局数","広義(在薬)薬局数","在総シェア"]
    for j,h in enumerate(heads,1): ws.cell(row=4,column=j,value=h)
    style_header(ws,4,len(heads))
    order2=sorted([(k,v) for k,v in agg.items() if v["tai"]>0],
                  key=lambda kv:(kv[1]["rid"]=="-",-kv[1]["tai"],conf_rank.get(kv[1]["conf"],3))); r=5
    for rank,(corp,info) in enumerate(order2,1):
        ws.cell(row=r,column=1,value=rank); ws.cell(row=r,column=2,value=corp)
        ws.cell(row=r,column=3,value=info["type"]); ws.cell(row=r,column=4,value=info["conf"])
        ws.cell(row=r,column=5,value=f"=COUNTIFS('10_ローデータ'!{CORP}:{CORP},B{r},'10_ローデータ'!{TAI}:{TAI},1)")
        ws.cell(row=r,column=6,value=f"=COUNTIFS('10_ローデータ'!{CORP}:{CORP},B{r},'10_ローデータ'!{HOM}:{HOM},1)")
        ws.cell(row=r,column=7,value=f"=E{r}/$E${5+len(order2)+1}"); ws.cell(row=r,column=7).number_format='0.0%'
        for c in range(1,8): ws.cell(row=r,column=c).font=NORM; ws.cell(row=r,column=c).border=BORD
        if info["rid"]!="-":
            for c in range(1,8): ws.cell(row=r,column=c).fill=GFILL
        r+=1
    ws.cell(row=r,column=2,value="合計（検算）").font=BOLD; ws.cell(row=r,column=5,value=f"=SUM(E5:E{r-1})").font=BOLD
    ws.cell(row=r+1,column=2,value="在総薬局 総数").font=NORM
    ws.cell(row=r+1,column=5,value=f"=COUNTIF('10_ローデータ'!{TAI}2:{TAI}{nrow},1)").font=NORM
    for j,w in enumerate([6,32,16,9,12,15,11],1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A5"

    path=f"{OUT}/福岡県_事業者集計.xlsx"; wb.save(path); return path,len(fk),len(order2)

if __name__=="__main__":
    rows,hist=build_rows()
    p1,n1,o1=build_fukuoka(rows,hist); print(f"福岡: {p1} 在宅薬局含む全{n1}件 在宅事業者{o1}")
    p2,n2,o2=build_kyushu(rows,hist); print(f"九州8県: {p2} 全{n2}件 在宅事業者{o2}")
